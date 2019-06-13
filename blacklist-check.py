import openpyxl
import os

mydir = input('Path: ')
os.chdir(mydir)

blacklist = []
checklist = []
exclusionList = []

def getblacklist():
    blacklist_workbook = openpyxl.load_workbook(input('Blacklist File Name: ') + '.xlsx')
    sheet = blacklist_workbook[input('Sheet Name:')]
    targetcol = int(input('Read Column: '))
    maxrange = input('Max Range: ')
    for i in range(targetcol, maxrange):
        blacklist.append(sheet.cell(row=i, column=2).value)

def getchecklist():
    checklist_workbook = openpyxl.load_workbook(input('Checklist File Name: ') + '.xlsx')
    cl_sheet = checklist_workbook[input('Sheet Name: ')]
    targetcol = int(input('Read Column: '))
    maxrange = input('Max Range: ')
    for i in range(targetcol, maxrange):
        checklist.append(cl_sheet.cell(row=i, column=5).value)

def blacklistcheck():
    for i in blacklist:
        if i in checklist and i is not None:
            exclusionList.append(i)

if __name__ == '__main__':
    getblacklist()
    getchecklist()
    blacklistcheck()
    print(exclusionList)
